import os
import sys
import comtypes.client
from HGAtools import *
from xldf import xldf

import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

#set the following flag to True to attach to an existing instance of the program
#otherwise a new instance of the program will be started
AttachToInstance = True

#set the following flag to True to manually specify the path to SAP2000.exe
#this allows for a connection to a version of SAP2000 other than the latest installation
#otherwise the latest installed version of SAP2000 will be launched
SpecifyPath = False


#create API helper object
helper = comtypes.client.CreateObject('SAP2000v1.Helper')
helper = helper.QueryInterface(comtypes.gen.SAP2000v1.cHelper)
mySapObject = helper.GetObject("CSI.SAP2000.API.SapObject") 


if AttachToInstance:
    #attach to a running instance of SAP2000
    try:
        #get the active SapObject
            mySapObject = helper.GetObject("CSI.SAP2000.API.SapObject") 

    except (OSError, comtypes.COMError):
        print("No running instance of the program found or failed to attach.")
        sys.exit(-1)


SapModel = mySapObject.SapModel

jd = Get_table('Joint Displacements', SapModel)
mpf = Get_table('Modal Periods And Frequencies', SapModel)


def select_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"),("Macro Enabled Excel Files", "*.xlsm")])
    return file_path

# def write_df_to_excel(df, file_path, sheet_name, insert_row, insert_column):

#     if not file_path:
#         print("No file selected.")
#         return
    
#     try:
#         book = load_workbook(file_path, keep_vba=True)  # Preserve macros
#         if sheet_name not in book.sheetnames:
#             print(f"Sheet '{sheet_name}' not found in {file_path}.")
#             return

#         # Use `ExcelWriter` without overwriting macros
#         with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#             writer.book = book  # Ensure the existing workbook (with macros) is used
#             df.to_excel(writer, sheet_name=sheet_name, startrow=insert_row, startcol=insert_column, index=False, header=True)
#             writer.close()  # Explicitly close writer
        
#         print(f"Data successfully written to {file_path} in sheet '{sheet_name}' at cell B2.")
#     except Exception as e:
#         print(f"Error: {e}")


# file_path = select_file()

# write_df_to_excel(jd.head(40), file_path, 'SAP Output - Mode Shapes', 1, 1)
# write_df_to_excel(mpf.head(40), file_path, 'SAP Output - Frequencies', 1, 1)


import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook


def write_df_to_xlsm(df, sheet_name):
    file_path = select_file()
    if not file_path:
        print("No file selected.")
        return
    
    #sheet_name = 'SAP Output - Mode Shapes'
    
    try:
        # Load the workbook with macros preserved
        book = load_workbook(file_path, keep_vba=True)

        if sheet_name not in book.sheetnames:
            print(f"Sheet '{sheet_name}' not found in {file_path}.")
            return
        
        sheet = book[sheet_name]

        # Convert DataFrame to list of lists (including column headers)
        data_with_headers = [df.columns.tolist()] + df.values.tolist()

        # Write data starting at cell B2 (which is row=2, col=2 in openpyxl)
        start_row, start_col = 2, 2  # B2 in Excel corresponds to (2,2)
        for i, row in enumerate(data_with_headers):
            for j, value in enumerate(row):
                sheet.cell(row=start_row + i, column=start_col + j, value=value)

        # Save the workbook
        book.save(file_path)
        print(f"Data successfully written to {file_path} in sheet '{sheet_name}' at cell B2.")

    except Exception as e:
        print(f"Error: {e}")

# Example DataFrame
# data = {'Column1': [1, 2, 3], 'Column2': ['A', 'B', 'C']}
# df = pd.DataFrame(data)

#write_df_to_xlsm(df)
write_df_to_xlsm(jd.head(40), 'SAP Output - Mode Shapes')
