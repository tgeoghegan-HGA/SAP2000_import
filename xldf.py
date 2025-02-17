
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import inspect


def xldf(df, df_string):

    desktop_path = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop') 
    path_to = desktop_path + f"\\{df_string}.xlsx"
    path_to = os.path.normpath(path_to)

    dataframe = df

    def export_dataframe_to_excel(dataframe, path_to):
        # Ensure the directory exists
        directory = os.path.dirname(path_to)
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Export DataFrame to Excel file
        with pd.ExcelWriter(path_to, engine='openpyxl') as writer:
            dataframe.to_excel(writer, index=True, sheet_name='Sheet1')

            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Apply center alignment to all cells
            for column in dataframe.columns:
                for cell in worksheet[get_column_letter(dataframe.columns.get_loc(column) + 1)]:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Autofit column width
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2  # Add a little extra width for padding

        # Save the workbook
        workbook.save(path_to)

        # Open Excel to the created file
        os.system(f'start excel "{path_to}"')

    export_dataframe_to_excel(dataframe, path_to)
